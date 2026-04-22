"""
expense_tracker.py  v4.0
========================
Local Expense Tracker — Universal Bank & Credit Card Statement Processor

FILENAME CONVENTION (required for Charged To auto-detection):
  [AccountPrefix]_[anything].csv
  Examples:
    AmexPlatinum_Jan2026.csv
    ChaseSapphire_Q12026.csv
    CapitalOneChecking_Mar2026.csv
    CapitalOneSparkBiz_2025Full.csv

ACCOUNT PREFIXES:
  AmexPlatinum, CapitalOneChecking, CapitalOneSparkBiz, ChaseSapphire,
  PayPalMC, DeltaBizPlat, ChaseAmazon, AmexHilton, ChaseMarriott, AmexDeltaPlat

CATEGORIES (your list):
  Cell Phone, Food, Internet, Medical, NJ Estimated Taxes, Rideshare,
  Subscriptions, Supplies, Travel, Toll, Rental Car Gas, Per Diem,
  Parking, Rental Car Parking, Payment for Tax Prep, Business Supplies,
  Rental Car, TV, Airport Pkg, Other

VENDOR MEMORY:
  Edit VENDOR_MEMORY dict below to permanently assign a category to a vendor.
  The script will auto-populate these on every run — no manual selection needed.
  After manually fixing a category in Excel, add that vendor here so it's
  remembered next time.

Requirements:
  pip install pandas openpyxl pdfplumber rapidfuzz
"""

import os
import sys
import re
import glob
import json
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from rapidfuzz import fuzz
    FUZZY_SUPPORT = True
except ImportError:
    FUZZY_SUPPORT = False


# ─────────────────────────────────────────────────────────────────────────────
# YOUR CATEGORY LIST
# This is the master list — all dropdowns and rules use exactly these values.
# Add or rename categories here and they'll flow through everywhere.
# ─────────────────────────────────────────────────────────────────────────────

CATEGORIES = [
    "Cell Phone",
    "Food",
    "Internet",
    "Medical",
    "NJ Estimated Taxes",
    "Rideshare",
    "Subscriptions",
    "Supplies",
    "Travel",
    "Toll",
    "Rental Car Gas",
    "Per Diem",
    "Parking",
    "Rental Car Parking",
    "Payment for Tax Prep",
    "Business Supplies",
    "Rental Car",
    "TV",
    "Airport Pkg",
    "Income",
    "Transfer",
    "CC Payment",
    "Other",
]


# ─────────────────────────────────────────────────────────────────────────────
# VENDOR MEMORY DATABASE
# Format: "keyword in vendor name (lowercase)": "Category"
# Rules are checked top-to-bottom; first match wins.
# After manually fixing a vendor's category in Excel, add it here so the
# script remembers it on every future run.
#
# HOW TO ADD: find the vendor name as it appears in your Excel Vendor column,
# take a distinctive word from it (lowercase), map it to your category.
# Example: you manually set "SPECTRUM" to "Internet" → add:
#   "spectrum": "Internet",
# ─────────────────────────────────────────────────────────────────────────────

VENDOR_MEMORY = {
    # ── Cell Phone ────────────────────────────────────────────────────────────
    "verizon wireless":         "Cell Phone",
    "t-mobile":                 "Cell Phone",
    "at&t wireless":            "Cell Phone",
    
    # ── Internet ──────────────────────────────────────────────────────────────
    "verizon fios":             "Internet",
    "optimum":                  "Internet",
    "xfinity":                  "Internet",
    "comcast":                  "Internet",
    "spectrum internet":        "Internet",

    # ── TV ────────────────────────────────────────────────────────────────────
    "spectrum tv":              "TV",
    "directv":                  "TV",
    "dish network":             "TV",
    "youtube tv":               "TV",
    "fubo":                     "TV",
    "sling":                    "TV",
    
    # ── Subscriptions ─────────────────────────────────────────────────────────
    "netflix":                  "Subscriptions",
    "hulu":                     "Subscriptions",
    "disney+":                  "Subscriptions",
    "hbo max":                  "Subscriptions",
    "max.com":                  "Subscriptions",
    "peacock":                  "Subscriptions",
    "paramount+":               "Subscriptions",
    "apple tv+":                "Subscriptions",
    "spotify":                  "Subscriptions",
    "apple music":              "Subscriptions",
    "amazon prime":             "Subscriptions",
    "prime video":              "Subscriptions",
    "adobe":                    "Subscriptions",
    "dropbox":                  "Subscriptions",
    "icloud":                   "Subscriptions",
    "google one":               "Subscriptions",
    "notion":                   "Subscriptions",
    "zoom":                     "Subscriptions",
    "slack":                    "Subscriptions",
    "openai":                   "Subscriptions",
    "anthropic":                "Subscriptions",
    "linkedin premium":         "Subscriptions",
    "sirius":                   "Subscriptions",

    # ── Food ──────────────────────────────────────────────────────────────────
    "whole foods":              "Food",
    "trader joe":               "Food",
    "wegmans":                  "Food",
    "shoprite":                 "Food",
    "stop & shop":              "Food",
    "aldi":                     "Food",
    "costco":                   "Food",
    "kroger":                   "Food",
    "safeway":                  "Food",
    "publix":                   "Food",
    "food bazaar":              "Food",
    "fairway":                  "Food",
    "starbucks":                "Food",
    "dunkin":                   "Food",
    "mcdonald":                 "Food",
    "burger king":              "Food",
    "chipotle":                 "Food",
    "panera":                   "Food",
    "doordash":                 "Food",
    "grubhub":                  "Food",
    "uber eats":                "Food",
    "instacart":                "Food",
    "seamless":                 "Food",
    
    # ── Rideshare ─────────────────────────────────────────────────────────────
    "uber":                     "Rideshare",
    "lyft":                     "Rideshare",
    "via ride":                 "Rideshare",
    "curb":                     "Rideshare",

    # ── Travel ────────────────────────────────────────────────────────────────
    "delta air":                "Travel",
    "united airlines":          "Travel",
    "american airlines":        "Travel",
    "southwest":                "Travel",
    "jetblue":                  "Travel",
    "spirit air":               "Travel",
    "alaska air":               "Travel",
    "british airways":          "Travel",
    "marriott":                 "Travel",
    "hilton":                   "Travel",
    "hyatt":                    "Travel",
    "airbnb":                   "Travel",
    "vrbo":                     "Travel",
    "amtrak":                   "Travel",

    # ── Rental Car ────────────────────────────────────────────────────────────
    "hertz":                    "Rental Car",
    "enterprise":               "Rental Car",
    "avis":                     "Rental Car",
    "budget rent":              "Rental Car",
    "national car":             "Rental Car",
    "alamo":                    "Rental Car",
    "dollar rent":              "Rental Car",

    # ── Rental Car Gas ────────────────────────────────────────────────────────
    "exxon":                    "Rental Car Gas",
    "shell":                    "Rental Car Gas",
    "chevron":                  "Rental Car Gas",
    "sunoco":                   "Rental Car Gas",
    "bp ":                      "Rental Car Gas",
    "marathon":                 "Rental Car Gas",
    "speedway":                 "Rental Car Gas",
    "wawa":                     "Rental Car Gas",
    "circle k":                 "Rental Car Gas",

    # ── Toll ──────────────────────────────────────────────────────────────────
    "e-zpass":                  "Toll",
    "ezpass":                   "Toll",
    "nj turnpike":              "Toll",
    "garden state pkwy":        "Toll",
    "port authority":           "Toll",

    # ── Parking ───────────────────────────────────────────────────────────────
    "spothero":                 "Parking",
    "parkwhiz":                 "Parking",
    "impark":                   "Parking",
    "laaz parking":             "Parking",

    # ── Airport Pkg ───────────────────────────────────────────────────────────
    "ewr parking":              "Airport Pkg",
    "jfk parking":              "Airport Pkg",
    "lga parking":              "Airport Pkg",
    "newark airport parking":   "Airport Pkg",

    # ── Rental Car Parking ────────────────────────────────────────────────────
    # Add rental car parking vendors here as you encounter them

    # ── Medical ───────────────────────────────────────────────────────────────
    "cvs":                      "Medical",
    "walgreen":                 "Medical",
    "rite aid":                 "Medical",
    "duane reade":              "Medical",
    "pharmacy":                 "Medical",

    # ── Supplies ──────────────────────────────────────────────────────────────
    "staples":                  "Supplies",
    "office depot":             "Supplies",
    "amazon":                   "Supplies",
    "target":                   "Supplies",
    "walmart":                  "Supplies",
    "home depot":               "Supplies",
    "lowes":                    "Supplies",

    # ── Business Supplies ─────────────────────────────────────────────────────
    "b&h photo":                "Business Supplies",
    "bhphotovideo":             "Business Supplies",
    "adorama":                  "Business Supplies",
    "filmtools":                "Business Supplies",
    "fedex":                    "Business Supplies",
    "ups store":                "Business Supplies",
    "usps":                     "Business Supplies",

    # ── Payment for Tax Prep ──────────────────────────────────────────────────
    "turbotax":                 "Payment for Tax Prep",
    "h&r block":                "Payment for Tax Prep",

    # ── NJ Estimated Taxes ────────────────────────────────────────────────────
    "nj div of taxation":       "NJ Estimated Taxes",
    "state of nj":              "NJ Estimated Taxes",
    "nj tax":                   "NJ Estimated Taxes",

    # ── Income ────────────────────────────────────────────────────────────────
    "payroll":                  "Income",
    "direct dep":               "Income",
    "adp":                      "Income",
    "gusto":                    "Income",
    "salary":                   "Income",
    "nbc payroll":              "Income",
    "nbcuniversal":             "Income",
    "tax refund":               "Income",
    "irs treas":                "Income",

    # ── Transfer ──────────────────────────────────────────────────────────────
    "zelle":                    "Transfer",
    "venmo":                    "Transfer",
    "cashapp":                  "Transfer",
    "wire transfer":            "Transfer",

    # ── CC Payment ────────────────────────────────────────────────────────────
    "payment thank you":        "CC Payment",
    "autopay":                  "CC Payment",
    "credit card payment":      "CC Payment",
    "bill payment":             "CC Payment",

    # ── ADD YOUR VENDORS BELOW ────────────────────────────────────────────────
    # Format: "vendor keyword (lowercase)": "Category from CATEGORIES list",
    "bjs":                      "Food",
    "vortex":                   "TV",
    "us mobile":                "Cell Phone",
}


# ─────────────────────────────────────────────────────────────────────────────
# ACCOUNT MAP
# Key   = filename prefix before underscore (case-insensitive, no spaces/dashes)
# Value = display name shown in "Charged To" column
# ─────────────────────────────────────────────────────────────────────────────

ACCOUNT_MAP = {
    "amexplatinum":        "AMEX PLATINUM",
    "capitalonechecking":  "CAPITAL ONE CHECKING",
    "capitalonesparkbiz":  "CAPITAL ONE SPARK BIZ",
    "chasesapphire":       "CHASE SAPPHIRE",
    "paypalmc":            "PAYPAL MC",
    "deltabizplat":        "DELTA BIZ PLAT",
    "chaseamazon":         "CHASE AMAZON",
    "amexhilton":          "AMEX HILTON",
    "chasemarriott":       "CHASE MARRIOTT",
    "amexdeltaplat":       "AMEX DELTA PLAT",
}

CHARGED_TO_OPTIONS = list(ACCOUNT_MAP.values()) + ["Unknown"]
DESKTOP_DIR = os.path.join(os.path.expanduser("~"), "Desktop")
STATEMENTS_DIR = os.path.join(DESKTOP_DIR, "Statements")
USER_VENDOR_MEMORY_PATH = os.path.join(DESKTOP_DIR, "expense_vendor_memory.json")
DEFAULT_OUTPUT_GLOB = "expense_report_*.xlsx"
VALIDATION_MAX_ROW = 5000

CATEGORY_LOOKUP = {str(category).strip().lower(): category for category in CATEGORIES}
CHARGED_TO_ALIASES = {
    "amex": "AMEX PLATINUM",
    "amex platinum": "AMEX PLATINUM",
    "american express platinum": "AMEX PLATINUM",
    "capital one": "CAPITAL ONE CHECKING",
    "capital one checking": "CAPITAL ONE CHECKING",
    "cap one": "CAPITAL ONE CHECKING",
    "cap one checking": "CAPITAL ONE CHECKING",
    "capital one spark": "CAPITAL ONE SPARK BIZ",
    "capital one spark biz": "CAPITAL ONE SPARK BIZ",
    "cap one spark": "CAPITAL ONE SPARK BIZ",
    "cap one spark biz": "CAPITAL ONE SPARK BIZ",
    "chase sapphire": "CHASE SAPPHIRE",
    "sapphire": "CHASE SAPPHIRE",
    "paypal": "PAYPAL MC",
    "paypal mc": "PAYPAL MC",
    "delta biz plat": "DELTA BIZ PLAT",
    "delta business platinum": "DELTA BIZ PLAT",
    "chase amazon": "CHASE AMAZON",
    "amazon chase": "CHASE AMAZON",
    "amazon prime chase": "CHASE AMAZON",
    "amex hilton": "AMEX HILTON",
    "hilton amex": "AMEX HILTON",
    "chase marriott": "CHASE MARRIOTT",
    "marriott chase": "CHASE MARRIOTT",
    "amex delta plat": "AMEX DELTA PLAT",
    "delta amex platinum": "AMEX DELTA PLAT",
    "unknown": "Unknown",
}


def normalize_account_key(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def normalize_memory_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def canonicalize_category(value):
    """Normalize loose category text to the exact allowed category value."""
    raw = str(value or "").strip()
    if not raw:
        return "Other"

    exact = CATEGORY_LOOKUP.get(raw.lower())
    if exact:
        return exact

    compact = re.sub(r"\s+", " ", raw.lower())
    exact = CATEGORY_LOOKUP.get(compact)
    if exact:
        return exact

    return "Other"


def canonicalize_charged_to(value):
    """Normalize loose account text to the exact allowed Charged To value."""
    raw = str(value or "").strip()
    if not raw:
        return "Unknown"

    if raw in CHARGED_TO_OPTIONS:
        return raw

    normalized = normalize_memory_key(raw)
    alias_match = CHARGED_TO_ALIASES.get(normalized)
    if alias_match:
        return alias_match

    normalized_key = normalize_account_key(raw)
    for key, display in ACCOUNT_MAP.items():
        if normalized_key == key or normalized_key in key or key in normalized_key:
            return display

    return "Unknown"


def normalize_category(value):
    return canonicalize_category(value)


def normalize_charged_to(value):
    return canonicalize_charged_to(value)


def detect_account(filepath):
    """
    Reads account name from filename prefix before first underscore.
    AmexPlatinum_Jan2026.csv → Amex Platinum
    Falls back to stem if no underscore found (still tries map lookup).
    """
    stem = Path(filepath).stem
    prefix = stem.split("_", 1)[0] if "_" in stem else stem

    for candidate in (prefix, stem):
        key = normalize_account_key(candidate)
        if key in ACCOUNT_MAP:
            return ACCOUNT_MAP[key]

    stem_key = normalize_account_key(stem)
    for key in sorted(ACCOUNT_MAP, key=len, reverse=True):
        if key in stem_key:
            return ACCOUNT_MAP[key]

    return "Unknown"


def load_user_vendor_memory(path=USER_VENDOR_MEMORY_PATH):
    """Load user-maintained vendor/category memory from Desktop JSON."""
    if not os.path.exists(path):
        return {}

    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception as exc:
        print(f"  [WARN] Could not read vendor memory file: {exc}")
        return {}

    if not isinstance(payload, dict):
        print("  [WARN] Vendor memory file is not a JSON object. Ignoring it.")
        return {}

    cleaned = {}
    for vendor, category in payload.items():
        key = normalize_memory_key(vendor)
        if key:
            cleaned[key] = canonicalize_category(category)
    return cleaned


def save_user_vendor_memory(memory, path=USER_VENDOR_MEMORY_PATH):
    """Persist learned vendor/category mappings to Desktop JSON."""
    cleaned = {}
    for vendor, category in memory.items():
        key = normalize_memory_key(vendor)
        if key:
            cleaned[key] = canonicalize_category(category)

    try:
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(dict(sorted(cleaned.items())), fh, indent=2, ensure_ascii=True)
    except Exception as exc:
        print(f"  [WARN] Could not save vendor memory file: {exc}")


def merge_vendor_memory(built_in_memory, user_memory):
    """Merge built-in memory with user memory. User mappings win on conflicts."""
    merged = {
        normalize_memory_key(vendor): canonicalize_category(category)
        for vendor, category in built_in_memory.items()
        if normalize_memory_key(vendor)
    }
    for vendor, category in user_memory.items():
        key = normalize_memory_key(vendor)
        if key:
            merged[key] = canonicalize_category(category)
    return merged


def find_previous_output_workbook(output_path=None, folder=DESKTOP_DIR):
    candidates = []
    for path in glob.glob(os.path.join(folder, DEFAULT_OUTPUT_GLOB)):
        if output_path and os.path.abspath(path) == os.path.abspath(output_path):
            continue
        candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def extract_manual_edits_from_previous_workbook(workbook_path):
    """
    Read Vendor + Category pairs from a previous Ledger sheet and convert them
    into learned vendor-memory entries.
    """
    learned = {}
    charged_to_seen = set()
    if not workbook_path or not os.path.exists(workbook_path):
        return learned

    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  [WARN] Could not read previous workbook for learning: {exc}")
        return learned

    try:
        if "Ledger" not in wb.sheetnames:
            return learned

        ws = wb["Ledger"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return learned

        headers = {str(value).strip(): idx for idx, value in enumerate(header_row)}
        vendor_idx = headers.get("Vendor")
        category_idx = headers.get("Category")
        charged_to_idx = headers.get("Charged To")
        if vendor_idx is None or category_idx is None:
            return learned

        for row in ws.iter_rows(min_row=2, values_only=True):
            vendor = clean_vendor_name(row[vendor_idx] or "")
            category = canonicalize_category(row[category_idx])
            vendor_key = normalize_memory_key(vendor)

            if charged_to_idx is not None:
                normalized_charged_to = canonicalize_charged_to(row[charged_to_idx])
                if normalized_charged_to != "Unknown":
                    charged_to_seen.add(normalized_charged_to)

            if not vendor_key or vendor == "Unknown Vendor":
                continue
            if category == "Other":
                continue
            if vendor_key in {"unknown vendor", "unknown", "nan"}:
                continue

            if vendor_key and category in CATEGORIES and category != "Other":
                learned[vendor_key] = category
    finally:
        wb.close()

    if learned:
        print(f"Learned {len(learned)} vendor mapping(s) from prior workbook.")
    if charged_to_seen:
        print(f"Normalized prior workbook Charged To values: {', '.join(sorted(charged_to_seen))}")

    return learned


# ─────────────────────────────────────────────────────────────────────────────
# UNIVERSAL COLUMN DETECTOR
# ─────────────────────────────────────────────────────────────────────────────

DATE_HINTS    = ["transaction date", "posting date", "post date", "posted date", "date"]
DESC_HINTS    = ["transaction description", "description", "payee", "merchant", "memo", "name", "details"]
AMOUNT_HINTS  = ["transaction amount", "amount"]
DEBIT_HINTS   = ["debit", "withdrawal", "charge"]
CREDIT_HINTS  = ["credit", "deposit", "payment"]
TYPE_HINTS    = ["transaction type", "type", "details"]


def find_col(columns, hints):
    cols_lower = {c: c.lower().strip() for c in columns}
    for hint in hints:
        for original, lower in cols_lower.items():
            if hint in lower:
                return original
    return None


def detect_amount_structure(columns):
    has_debit  = find_col(columns, DEBIT_HINTS) is not None
    has_credit = find_col(columns, CREDIT_HINTS) is not None
    return "split" if (has_debit and has_credit) else "single"


def parse_amount(val):
    if pd.isna(val) or str(val).strip() in ("", "-", "–"):
        return 0.0
    s = str(val).replace("$", "").replace(",", "").replace(" ", "").strip()
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def format_amount(val):
    """Format as $1,234.56 or -$1,234.56"""
    if val < 0:
        return f"-${abs(val):,.2f}"
    return f"${val:,.2f}"


def normalize_file(filepath):
    """Universal CSV normalizer — works with any bank's export format."""
    print(f"  Loading: {Path(filepath).name}")

    df = None
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            df = pd.read_csv(filepath, encoding=enc, on_bad_lines="skip",
                             skip_blank_lines=True)
            df.columns = df.columns.str.strip()
            df = df.dropna(axis=1, how="all")
            if len(df.columns) >= 2:
                break
        except Exception:
            continue

    if df is None or df.empty:
        raise ValueError("Could not read file or file is empty")

    df      = df.dropna(how="all").reset_index(drop=True)
    columns = df.columns.tolist()
    account = detect_account(filepath)
    print(f"    Detected account: {account}")
    out     = pd.DataFrame()
    out["Charged_To"] = account

    # Date
    date_col = find_col(columns, DATE_HINTS)
    if not date_col:
        raise ValueError("Could not find a Date column")
    out["Date"] = pd.to_datetime(df[date_col], errors="coerce")

    # Description
    desc_col = find_col(columns, DESC_HINTS)
    if not desc_col:
        raise ValueError("Could not find a Description/Vendor column")
    out["Vendor_raw"] = df[desc_col].astype(str).str.strip()

    # Amount
    structure = detect_amount_structure(columns)
    if structure == "split":
        debit_col  = find_col(columns, DEBIT_HINTS)
        credit_col = find_col(columns, CREDIT_HINTS)
        debits  = df[debit_col].apply(parse_amount)  if debit_col  else pd.Series([0.0]*len(df))
        credits = df[credit_col].apply(parse_amount) if credit_col else pd.Series([0.0]*len(df))
        out["Amount_raw"] = credits - debits
    else:
        amt_col = find_col(columns, AMOUNT_HINTS)
        if not amt_col:
            for c in reversed(columns):
                try:
                    test = df[c].apply(parse_amount)
                    if test.abs().sum() > 0:
                        amt_col = c
                        break
                except Exception:
                    continue
        if not amt_col:
            raise ValueError("Could not find an Amount column")

        amounts  = df[amt_col].apply(parse_amount)
        type_col = find_col(columns, TYPE_HINTS)
        if type_col:
            types         = df[type_col].astype(str).str.lower()
            has_type_info = types.str.contains(
                "debit|credit|deposit|withdrawal", na=False).any()
            if has_type_info:
                def signed_amount(row):
                    t = str(row[type_col]).lower()
                    a = parse_amount(row[amt_col])
                    if any(k in t for k in ["debit", "withdrawal", "charge", "purchase"]):
                        return -abs(a)
                    elif any(k in t for k in ["credit", "deposit", "payment", "return"]):
                        return abs(a)
                    return a
                out["Amount_raw"] = df.apply(signed_amount, axis=1)
            else:
                out["Amount_raw"] = amounts
        else:
            out["Amount_raw"] = amounts

    out = out.dropna(subset=["Date"])
    out = out[out["Vendor_raw"].str.strip().str.len() > 0]
    out = out[out["Amount_raw"] != 0]

    print(f"    -> {account} | {len(out)} transactions")
    return out


def normalize_pdf(filepath):
    """Extract transactions from PDF bank statements."""
    if not PDF_SUPPORT:
        print(f"  [SKIP] PDF support requires pdfplumber: pip install pdfplumber")
        return pd.DataFrame()

    print(f"  Loading PDF: {Path(filepath).name}")
    account = detect_account(filepath)
    print(f"    Detected account: {account}")
    rows    = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row and any(cell for cell in row if cell):
                            rows.append([str(c).strip() if c else "" for c in row])
            else:
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        rows.append([line])

    if not rows:
        print(f"  [WARN] No content in PDF. Export as CSV for best results.")
        return pd.DataFrame()

    if len(rows[0]) >= 3:
        try:
            import tempfile, csv
            df_raw = pd.DataFrame(rows[1:], columns=[str(c).strip() for c in rows[0]])
            df_raw = df_raw.dropna(how="all")
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".csv", delete=False, newline="",
                prefix=Path(filepath).stem + "_"
            ) as tmp:
                writer = csv.writer(tmp)
                writer.writerow(df_raw.columns.tolist())
                for _, r in df_raw.iterrows():
                    writer.writerow(r.tolist())
                tmp_path = tmp.name
            result = normalize_file(tmp_path)
            os.unlink(tmp_path)
            return result
        except Exception:
            pass

    date_pat   = re.compile(r"\b(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})\b")
    amount_pat = re.compile(r"\(?\$?([\d,]+\.\d{2})\)?")
    records    = []

    for row in rows:
        line     = " ".join(row)
        date_m   = date_pat.search(line)
        amount_m = amount_pat.findall(line)
        if date_m and amount_m:
            try:
                date_val   = pd.to_datetime(date_m.group(1), errors="coerce")
                amount     = parse_amount(amount_m[-1])
                desc_start = date_m.end()
                desc_end   = line.rfind(amount_m[-1])
                desc       = line[desc_start:desc_end].strip(" -$,")
                if desc and date_val is not pd.NaT:
                    records.append({
                        "Date": date_val, "Vendor_raw": desc,
                        "Amount_raw": -amount, "Charged_To": account,
                    })
            except Exception:
                continue

    if not records:
        print(f"  [WARN] Could not parse PDF. Export as CSV for best results.")
        return pd.DataFrame()

    out = pd.DataFrame(records)
    print(f"    -> {account} | {len(out)} transactions (PDF)")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFIER
# Uses VENDOR_MEMORY first (exact/substring match), then fuzzy fallback.
# ─────────────────────────────────────────────────────────────────────────────

def classify_vendor(vendor_clean, vendor_memory, vendor_raw=""):
    """
    Match a cleaned vendor name against merged vendor memory. Falls back to the
    raw description only if the cleaned value does not produce a match.
    """
    candidates = [normalize_memory_key(vendor_clean)]
    raw_candidate = normalize_memory_key(vendor_raw)
    if raw_candidate and raw_candidate not in candidates:
        candidates.append(raw_candidate)

    memory_items = sorted(vendor_memory.items(), key=lambda item: len(item[0]), reverse=True)

    for candidate in candidates:
        if not candidate:
            continue

        exact_match = vendor_memory.get(candidate)
        if exact_match:
            return exact_match

        for keyword, category in memory_items:
            if keyword and keyword in candidate:
                return category

    if FUZZY_SUPPORT:
        best_score, best_cat = 0, None
        primary_candidate = candidates[0]
        if primary_candidate:
            for keyword, category in memory_items:
                score = fuzz.partial_ratio(keyword, primary_candidate)
                if score > best_score:
                    best_score, best_cat = score, category
            if best_score >= 85 and best_cat:
                return best_cat

    return "Other"


def detect_transaction_type(amount, vendor):
    v = vendor.lower()
    if re.search(r"\btransfer\b|\bzelle\b|\bvenmo\b|\bwire\b", v):
        return "Transfer"
    if re.search(r"\bfee\b|\bpenalty\b", v):
        return "Fee"
    if re.search(r"refund|return|reversal|credit adj", v):
        return "Refund"
    if amount > 0:
        return "Credit"
    return "Debit"


def clean_vendor_name(raw):
    """Strip bank noise prefixes from transaction descriptions."""
    prefixes = [
        r"^purchase authorized on \d{2}/\d{2} ",
        r"^pos purchase\s*[-]?\s*",
        r"^preauthorized (withdrawal|payment|debit) (to|from)?\s*",
        r"^withdrawal from\s*",
        r"^deposit from\s*",
        r"^recurring (payment|charge|debit)\s*[-]?\s*",
        r"^ach (payment|debit|credit)\s*[-]?\s*",
        r"^checkcard\s*",
        r"^\d{2}/\d{2}\s+",
        r"\s+#\d{4,}.*$",
        r"\s+\d{3,}-\d{3,}.*$",
    ]
    cleaned = str(raw or "").strip()
    for p in prefixes:
        cleaned = re.sub(p, "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned[:60] if cleaned else "Unknown Vendor"


def classify_all(df, vendor_memory):
    """Apply all classification and formatting to the combined DataFrame."""
    df = df.copy()
    df["Vendor_raw"] = df["Vendor_raw"].fillna("").astype(str).str.strip()
    df["Vendor"] = df["Vendor_raw"].apply(clean_vendor_name)
    df["Category"] = df.apply(
        lambda r: classify_vendor(r["Vendor"], vendor_memory, r["Vendor_raw"]),
        axis=1,
    )
    df["Category"] = df["Category"].apply(normalize_category)
    df["Charged_To"] = df["Charged_To"].apply(normalize_charged_to)
    if "Category" in df.columns:
        df["Category"] = df["Category"].apply(canonicalize_category)
    if "Charged_To" in df.columns:
        df["Charged_To"] = df["Charged_To"].apply(canonicalize_charged_to)
    df["Show"] = df.get("Show", "").fillna("") if "Show" in df.columns else ""
    df["Notes"] = df.get("Notes", "").fillna("") if "Notes" in df.columns else ""
    df["Amount_raw"] = pd.to_numeric(df["Amount_raw"], errors="coerce").fillna(0).round(2)
    df["Amount_fmt"] = df["Amount_raw"].apply(format_amount)
    df["Transaction_Type"] = df.apply(
        lambda r: detect_transaction_type(r["Amount_raw"], r["Vendor"]),
        axis=1,
    )
    return df


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

COLORS = {
    "header":    "1B3A4B",
    "header_fg": "FFFFFF",
    "alt_row":   "F4F7FA",
    "credit":    "E6F4EA",
    "transfer":  "EEF2FF",
    "other":     "FFF8E1",
}

def hdr_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def row_fill(hex_color):  return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0D5DD")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=10):
    return Font(bold=True, color=COLORS["header_fg"], size=size)

def apply_header(ws, headers):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill(COLORS["header"])
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 28


def safe_cell_value(value, default=""):
    if pd.isna(value) or value is None:
        return default
    return value


def ensure_lists_sheet(wb):
    """Create or refresh the hidden _Lists sheet used by Excel validations."""
    hidden_name = "_Lists"
    if hidden_name in wb.sheetnames:
        ws = wb[hidden_name]
        ws.delete_rows(1, ws.max_row or 1)
    else:
        ws = wb.create_sheet(hidden_name)

    for row, category in enumerate(CATEGORIES, start=1):
        ws.cell(row=row, column=1, value=category)

    for row, charged_to in enumerate(CHARGED_TO_OPTIONS, start=1):
        ws.cell(row=row, column=2, value=charged_to)

    ws.sheet_state = "hidden"
    return ws


def create_named_range(wb, name, attr_text):
    """Create or replace a workbook-scoped named range for validation lists."""
    try:
        del wb.defined_names[name]
    except Exception:
        pass
    wb.defined_names[name] = DefinedName(name=name, attr_text=attr_text)


def ensure_validation_named_ranges(wb):
    """Create the workbook-level named ranges used by Ledger validation."""
    ensure_lists_sheet(wb)
    create_named_range(wb, "CategoryList", f"'_Lists'!$A$1:$A${len(CATEGORIES)}")
    create_named_range(wb, "ChargedToList", f"'_Lists'!$B$1:$B${len(CHARGED_TO_OPTIONS)}")
    print("Created named ranges: CategoryList, ChargedToList")


def add_list_validation(ws, start_row, end_row, col_idx, list_name, prompt_title, prompt_text,
                        error_title, error_text):
    """
    Add data validation dropdowns to cells. Header arrows are filters; these
    are the actual editable cell dropdowns.
    """
    if end_row < start_row:
        return

    col_letter = get_column_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1=f"={list_name}",
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle=error_title,
        error=error_text,
        showInputMessage=True,
        promptTitle=prompt_title,
        prompt=prompt_text,
    )
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv)


def _obsolete_build_category_dropdown(ws, start_row, end_row, col):
    """
    Add a dropdown using a hidden sheet reference — works across all Excel versions.
    """
    wb = ws.parent
    hidden_name = "_Categories"
    if hidden_name not in wb.sheetnames:
        hidden_ws = wb.create_sheet(hidden_name)
        for i, cat in enumerate(CATEGORIES, start=1):
            hidden_ws.cell(row=i, column=1, value=cat)
        hidden_ws.sheet_state = "hidden"
    ref        = f"_Categories!$A$1:$A${len(CATEGORIES)}"
    dv         = DataValidation(type="list", formula1=ref,
                                allow_blank=True, showDropDown=False)
    col_letter = get_column_letter(col)
    dv.sqref   = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def write_ledger(ws, df):
    """
    Main ledger tab.
    Columns: Date | Vendor | Category | Charged To | Amount | Show | Notes
    Category column has a dropdown on every data row.
    """
    headers = [
        "Date", "Vendor", "Category", "Charged To",
        "Amount", "Transaction Type", "Show", "Notes"
    ]
    apply_header(ws, headers)
    ensure_validation_named_ranges(ws.parent)

    sorted_df = df.sort_values("Date").reset_index(drop=True)
    total_rows = len(sorted_df)
    cat_col_idx = headers.index("Category") + 1
    charged_to_col_idx = headers.index("Charged To") + 1
    validation_end_row = max(total_rows + 1, VALIDATION_MAX_ROW)

    for i, (_, row) in enumerate(sorted_df.iterrows(), start=2):
        amt_raw = float(safe_cell_value(row.get("Amount_raw", 0), 0) or 0)
        tx_type = safe_cell_value(row.get("Transaction_Type", ""), "")
        cat = canonicalize_category(safe_cell_value(row.get("Category", "Other"), "Other"))
        charged_to = canonicalize_charged_to(
            safe_cell_value(row.get("Charged_To", "Unknown"), "Unknown")
        )
        vendor = safe_cell_value(row.get("Vendor", ""), "")
        show_value = safe_cell_value(row.get("Show", ""), "")
        notes_value = safe_cell_value(row.get("Notes", ""), "")

        vals = [
            row["Date"].date() if pd.notnull(row.get("Date")) else "",
            vendor,
            cat,
            charged_to,
            safe_cell_value(row.get("Amount_fmt", ""), ""),
            tx_type,
            show_value,
            notes_value,
        ]
        ws.append(vals)

        # Row shading
        if tx_type == "Transfer":
            fill_color = COLORS["transfer"]
        elif amt_raw > 0:
            fill_color = COLORS["credit"]
        elif cat == "Other":
            fill_color = COLORS["other"]
        elif i % 2 == 0:
            fill_color = COLORS["alt_row"]
        else:
            fill_color = None

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin_border()
            if fill_color:
                cell.fill = row_fill(fill_color)

        ws.cell(row=i, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=1).number_format = "MM/DD/YYYY"

    # Header arrows are filters. Cell dropdowns below are data validation.
    add_list_validation(
        ws,
        2,
        validation_end_row,
        cat_col_idx,
        "CategoryList",
        "Category",
        "Pick a category from the dropdown list.",
        "Invalid Category",
        "Select a valid Category from the dropdown list.",
    )
    add_list_validation(
        ws,
        2,
        validation_end_row,
        charged_to_col_idx,
        "ChargedToList",
        "Charged To",
        "Pick a Charged To value from the dropdown list.",
        "Invalid Charged To",
        "Select a valid Charged To value from the dropdown list.",
    )
    print(
        f"Applied data validation to Ledger Category and Charged To columns through row {validation_end_row}."
    )

    # Column widths
    widths = [12, 40, 22, 22, 14, 16, 20, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_monthly_summary(ws, df):
    """Spending by Category × Month."""
    debits = df[df["Amount_raw"] < 0].copy()
    if debits.empty:
        ws.append(["No debit transactions found."])
        return

    debits["Month"] = debits["Date"].dt.to_period("M").astype(str)
    pivot = debits.pivot_table(
        index="Category", columns="Month",
        values="Amount_raw", aggfunc="sum"
    ).fillna(0).abs()
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False)

    months = list(pivot.columns)
    apply_header(ws, ["Category"] + months)

    for i, (cat, row_data) in enumerate(pivot.iterrows(), start=2):
        ws.append([cat] + [round(v, 2) for v in row_data.values])
        for col in range(2, len(months) + 2):
            c = ws.cell(row=i, column=col)
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()
        ws.cell(row=i, column=1).border = thin_border()
        if i % 2 == 0:
            for col in range(1, len(months) + 2):
                ws.cell(row=i, column=col).fill = row_fill(COLORS["alt_row"])

    ws.column_dimensions["A"].width = 28
    for j in range(2, len(months) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:A1"


def write_by_account(ws, df):
    """Spending breakdown by account."""
    debits = df[df["Amount_raw"] < 0].copy()
    if debits.empty:
        ws.append(["No debit transactions found."])
        return

    apply_header(ws, ["Charged To", "Category", "Total Spent"])
    by_acct   = debits.groupby(["Charged_To", "Category"])["Amount_raw"].sum().abs()
    prev_acct = None

    for i, ((acct, cat), total) in enumerate(by_acct.items(), start=2):
        ws.append([acct if acct != prev_acct else "", cat, round(total, 2)])
        prev_acct = acct
        ws.cell(i, 3).number_format = '"$"#,##0.00'
        ws.cell(i, 3).alignment = Alignment(horizontal="right")
        for col in range(1, 4):
            ws.cell(i, col).border = thin_border()
        if i % 2 == 0:
            for col in range(1, 4):
                ws.cell(i, col).fill = row_fill(COLORS["alt_row"])

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A2"


def write_show_report(ws, df):
    """Expenses by Show/Project."""
    ws.cell(1, 1, "EXPENSE REPORT BY SHOW / PROJECT")
    ws.cell(1, 1).font = Font(bold=True, size=13, color=COLORS["header"])
    ws.cell(2, 1, f"Generated: {datetime.now().strftime('%B %d, %Y')}")
    ws.append([])

    show_df = df[(df["Show"] != "") & (df["Amount_raw"] < 0)].copy()
    if show_df.empty:
        ws.cell(4, 1, "No show-tagged transactions yet.")
        ws.cell(5, 1, "To tag shows: edit the SHOW_KEYWORDS dict in the script.")
        return

    by_show      = show_df.groupby(["Show", "Category"])["Amount_raw"].sum().abs()
    current_show = None
    row_num      = 4

    for (show, cat), total in by_show.items():
        if show != current_show:
            ws.cell(row_num, 1, f"SHOW: {show}")
            ws.cell(row_num, 1).font = Font(bold=True, size=11, color=COLORS["header"])
            ws.cell(row_num, 1).fill = hdr_fill("D1E8F0")
            row_num += 1
            current_show = show
        ws.cell(row_num, 2, cat)
        ws.cell(row_num, 3, round(total, 2))
        ws.cell(row_num, 3).number_format = '"$"#,##0.00'
        row_num += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def get_files_from_picker():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root  = tk.Tk()
        root.withdraw()
        files = filedialog.askopenfilenames(
            title="Select statement files (CSV or PDF)",
            filetypes=[
                ("CSV and PDF", "*.csv *.pdf"),
                ("CSV files",   "*.csv"),
                ("PDF files",   "*.pdf"),
                ("All files",   "*.*"),
            ]
        )
        root.destroy()
        return list(files)
    except Exception:
        return []


def main():
    parser = argparse.ArgumentParser(description="Expense Tracker v4.0")
    parser.add_argument("files",    nargs="*", help="CSV or PDF statement files")
    parser.add_argument("--output", default=None, help="Output Excel filename")
    args = parser.parse_args()
    files = list(args.files)
    outfile = args.output or f"expense_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    out_path = os.path.join(DESKTOP_DIR, outfile)

    if not files:
        files = (glob.glob(os.path.join(STATEMENTS_DIR, "*.csv")) +
                 glob.glob(os.path.join(STATEMENTS_DIR, "*.pdf")))
        if files:
            print(f"Auto-detected {len(files)} file(s) in Desktop\\Statements")
        else:
            print("No files in Desktop\\Statements — opening file picker...")
            files = get_files_from_picker()

    if not files:
        print("No files selected. Exiting.")
        sys.exit(0)

    print(f"\n{'='*60}")
    print(f"  EXPENSE TRACKER v4.0 — processing {len(files)} file(s)")
    print(f"{'='*60}\n")

    all_dfs = []
    for f in files:
        ext = Path(f).suffix.lower()
        try:
            df = normalize_pdf(f) if ext == ".pdf" else normalize_file(f)
            if not df.empty:
                all_dfs.append(df)
        except Exception as e:
            print(f"  [ERROR] {Path(f).name}: {e}")

    if not all_dfs:
        print("\nNo valid data found. Confirm files are bank-exported CSVs.")
        sys.exit(1)

    user_vendor_memory = load_user_vendor_memory()
    previous_workbook = find_previous_output_workbook(output_path=out_path)
    learned_memory = extract_manual_edits_from_previous_workbook(previous_workbook)
    if learned_memory:
        user_vendor_memory.update(learned_memory)
        save_user_vendor_memory(user_vendor_memory)
        print(f"Loaded {len(learned_memory)} learned vendor mapping(s) from prior workbook.")

    vendor_memory = merge_vendor_memory(VENDOR_MEMORY, user_vendor_memory)
    combined = pd.concat(all_dfs, ignore_index=True)
    combined = combined.dropna(subset=["Date"])
    combined = combined.sort_values("Date").reset_index(drop=True)
    combined = classify_all(combined, vendor_memory)
    print("Unique Charged To values:", ", ".join(sorted(combined["Charged_To"].dropna().unique())))
    print("Unique Category values:", ", ".join(sorted(combined["Category"].dropna().unique())))

    total       = len(combined)
    other_count = (combined["Category"] == "Other").sum()
    spending    = combined[combined["Amount_raw"] < 0]["Amount_raw"].sum()
    income      = combined[combined["Amount_raw"] > 0]["Amount_raw"].sum()

    print(f"\n{'─'*60}")
    print(f"  Transactions processed : {total:,}")
    print(f"  Auto-categorized       : {total - other_count:,}")
    print(f"  Needs manual category  : {other_count:,}  ← use dropdown in Excel")
    print(f"  Total spend            : {format_amount(abs(spending))}")
    print(f"  Total income/credits   : {format_amount(income)}")
    print(f"{'─'*60}")
    print(f"\n  TIP: Manual Ledger edits can be learned into:")
    print(f"  {USER_VENDOR_MEMORY_PATH}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Writing Excel workbook...")
    write_ledger(         wb.create_sheet("Ledger"),          combined)
    write_monthly_summary(wb.create_sheet("Monthly Summary"), combined)
    write_by_account(     wb.create_sheet("By Account"),      combined)
    write_show_report(    wb.create_sheet("By Show"),         combined)

    wb.save(out_path)
    print(f"\n  Done! Saved to: {out_path}\n")

    if sys.platform == "win32":
        os.startfile(out_path)


if __name__ == "__main__":
    main()
